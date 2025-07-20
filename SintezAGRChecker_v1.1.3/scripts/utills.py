
import os
import struct
import imghdr
import math
import csv
import traceback
import sys
import subprocess
import site

import bpy
import bmesh
import mathutils
import addon_utils

from . import check_highpoly_lowpoly
from . import utills
from . import ui_utills
from . import os_utils
from . import logger


CHECK_STATE_ITEMS = [
    "undefined",
    "verified",
    "failed"
]

class UdimSet():
    def __init__(self):
        self.diffuse_set = []
        self.erm_set = []
        self.normal_set = []
        self.resolutions_by_number = dict()
        self.sets_count = 0
    
    @property
    def collection(self):
        result = []
        for i in range(len(self.diffuse_set)):
            result.append((self.diffuse_set[i], self.erm_set[i], self.normal_set[i]))
        return result

def create_udim_sets(root):
    udim_set = UdimSet()
    files = [f for f in os.listdir(root) if ".png" in f]
    if not files:
        return udim_set

    count = max([udim_number(f) for f in files]) - 1000
    udim_set.sets_count = count
    udim_set.diffuse_set = [None for _ in range(count)]
    udim_set.erm_set = [None for _ in range(count)]
    udim_set.normal_set = [None for _ in range(count)]

    for file in files:
        ind = udim_number(file) - 1000 - 1
        if "diffuse" in file.lower() or "basecolor" in file.lower().replace(" ", ""):
            # udim_set.diffuse_set.append(file)
            udim_set.diffuse_set[ind] = file
        elif "_erm_" in file.lower():
            udim_set.erm_set[ind] = file
        elif "normal" in file.lower():
            udim_set.normal_set[ind] = file

    for set in udim_set.collection:
        max_res = 0
        for file in set:
            if file == None:
                continue
            num = udim_number(file)
            res = get_image_size(root + "\\" + file)[0]
            if res > max_res:
                max_res = res
        if max_res != 0:
            udim_set.resolutions_by_number[num] = max_res
    return udim_set

def udim_number(name):
    return int(name[-8:-4])

def get_image_size(fname):
    '''Determine the image type of fhandle and return its size.
    from draco'''
    with open(fname, 'rb') as fhandle:
        head = fhandle.read(24)
        if len(head) != 24:
            return
        if imghdr.what(fname) == 'png':
            check = struct.unpack('>i', head[4:8])[0]
            if check != 0x0d0a1a0a:
                return
            width, height = struct.unpack('>ii', head[16:24])
        else:
            return
        return width, height

def uv_to_udim_number(x, y):
    return 1000 + math.floor(y) * 10 + math.ceil(x)

def _td_errors_by_udim(obj, udim_resolutions, highpoly, td_min, td_max):
    # udim_resolutions = {1001: 4096, 1002: 256, 1003: 2048}

    out_udim_face_indices = []
    # td_errors_less = dict()
    # td_errors_greater = dict()
    # not_uvmap_errors = []
    # calculated_obj_td_list = []

    td_less_indices = []
    td_greater_indices = []

    bpy.ops.object.mode_set(mode='OBJECT')
    bpy.context.view_layer.objects.active = obj
    bpy.ops.object.mode_set(mode='EDIT')
    bm = bmesh.from_edit_mesh(obj.data)
    bmesh.ops.triangulate(bm, faces=bm.faces[:])
    bm.faces.ensure_lookup_table()

    face_count = len(obj.data.polygons)
    udims_used = set()

    vector_up = mathutils.Vector((0, 0, 1.0))

    # test_counter = 0
    for x in range(0, face_count):
        # bpy.context.scene.agr_scene_properties.progress = x / face_count
        if not highpoly and round(math.degrees(obj.data.polygons[x].normal.angle(vector_up))) == 90:
            continue
        area = 0
        loops = []
        try:
            for loop in bm.faces[x].loops:
                loops.append(loop[bm.loops.layers.uv.active].uv)
        except:
            logger.add("cant find BMLayerItem in bm.loops.layers.uv.active")
            # not_uvmap_errors.append(f"cant find BMLayerItem in bm.loops.layers.uv.active={bm.loops.layers.uv.active}, obj={obj.name}")
            continue
        loops_count = len(loops)
        a = loops_count - 1

        if highpoly:
            udim_nums = [uv_to_udim_number(loops[i].x, loops[i].y) for i in range(loops_count)]
            [udims_used.add(num) for num in udim_nums]
            # check if all points fall into one tile and tile exist in udims
            if not all(num == udim_nums[0] for num in udim_nums) or udim_nums[0] not in udim_resolutions.keys():
                out_udim_face_indices.append(x)
                continue
            udim_num = udim_nums[0]
        else:
            udim_num = 1001

        largest_side = udim_resolutions[udim_num]
        if largest_side == None:
            continue
        if largest_side <= 256:
            continue
        
        for b in range(0, loops_count):
            area += (loops[a].x + loops[b].x) * (loops[a].y - loops[b].y)
            a = b
        
        area = abs(0.5 * area)
        gm_area = obj.data.polygons[x].area

        if gm_area > 0 and area > 0:
            aspect_ratio = 1
            texel_density = ((largest_side / math.sqrt(aspect_ratio)) * math.sqrt(area))/(math.sqrt(gm_area)*100) / bpy.context.scene.unit_settings.scale_length
        else:
            texel_density = 0.0001

        texel_density = texel_density * 100
        # test_counter += 1
        # if test_counter % 100 == 0 and test_counter < 3000:
        #     logger.add(f"td={texel_density}")

        if highpoly:
            # texel_min, texel_max = 512, 1706
            texel_min, texel_max = td_min, td_max
        else:
            texel_min, texel_max = td_min, td_max
        if texel_density < texel_min:
            td_less_indices.append(x)
            # if udim_num not in td_errors_less.keys():
            #     td_errors_less[udim_num] = [largest_side, [], []]
            # td_errors_less[udim_num][1].append(round(texel_density, 2))
            # td_errors_less[udim_num][2].append(round(gm_area, 2))
        elif texel_density > texel_max:
            td_greater_indices.append(x)
            # if udim_num not in td_errors_greater.keys():
            #     td_errors_greater[udim_num] = [largest_side, [], []]
            # td_errors_greater[udim_num][1].append(round(texel_density, 2))
            # td_errors_greater[udim_num][2].append(round(gm_area, 2))

        # calculated_obj_td_list.append(texel_density)

    # udim_used_errors = []
    # for key in udim_resolutions.keys():
    #     if key not in udims_used:
    #         udim_used_errors.append(f"Неиспользуемая текстура, UDIM {key}")

    bpy.ops.object.mode_set(mode='OBJECT')

    return td_less_indices, td_greater_indices, out_udim_face_indices

def _td_errors(obj, tex_size, td_min, td_max):
    td_less_indices = []
    td_greater_indices = []

    bpy.ops.object.mode_set(mode='OBJECT')
    bpy.context.view_layer.objects.active = obj
    bpy.ops.object.transform_apply(location=False, rotation=False, scale=True, properties=False)
    bpy.ops.object.mode_set(mode='EDIT')
    bm = bmesh.from_edit_mesh(obj.data)
    bmesh.ops.triangulate(bm, faces=bm.faces[:])
    bm.faces.ensure_lookup_table()

    face_count = len(obj.data.polygons)

    for x in range(0, face_count):
        area = 0
        loops = []
        try:
            for loop in bm.faces[x].loops:
                loops.append(loop[bm.loops.layers.uv.active].uv)
        except:
            logger.add("cant find BMLayerItem in bm.loops.layers.uv.active")
            # not_uvmap_errors.append(f"cant find BMLayerItem in bm.loops.layers.uv.active={bm.loops.layers.uv.active}, obj={obj.name}")
            continue
        loops_count = len(loops)
        a = loops_count - 1
        largest_side = tex_size
        
        for b in range(0, loops_count):
            area += (loops[a].x + loops[b].x) * (loops[a].y - loops[b].y)
            a = b
        
        area = abs(0.5 * area)
        gm_area = obj.data.polygons[x].area

        if gm_area > 0 and area > 0:
            aspect_ratio = 1
            texel_density = ((largest_side / math.sqrt(aspect_ratio)) * math.sqrt(area))/(math.sqrt(gm_area)*100) / bpy.context.scene.unit_settings.scale_length
        else:
            texel_density = 0.0001

        texel_density = texel_density * 100

        texel_min, texel_max = td_min, td_max
        if texel_density < texel_min:
            td_less_indices.append(x)
        elif texel_density > texel_max:
            td_greater_indices.append(x)

    bpy.ops.object.mode_set(mode='OBJECT')

    return td_less_indices, td_greater_indices

def ensure_import_openpyxl():
    try:
        user_site_pkgs = site.getusersitepackages()
        if user_site_pkgs not in sys.path:
            sys.path.append(user_site_pkgs)
        import openpyxl
        return True
    except Exception as e:
        try:
            logger.add('installing package openpyxl')
            py_exec = sys.executable
            subprocess.call([str(py_exec), "-m", "ensurepip", "--user"])
            subprocess.call([str(py_exec), "-m", "pip", "install", "--upgrade", "pip"])
            subprocess.run([str(py_exec), "-m", "pip", "install", "openpyxl"])
            import openpyxl
            return True
        
        except Exception as e:
            logger.add_error(e, f"Не удалось установить библиотеку openpyxl (для чтения excel файлов)")
            return False

def get_lowpoly_codes_by_street(street):
    path = os_utils.get_documents_dir()
    
    if ensure_import_openpyxl():
        from openpyxl import load_workbook
        
        file_name = "lp_codes.xlsx"

        logger.add(f"Path to lowpoly codes {os.path.join(path, file_name)}")

        workbook = load_workbook(filename=os.path.join(path, file_name))
        sheet = workbook.active
        codes = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if street.lower() in str(row[1]).lower():
                codes.append(f"{row[1]} : {row[6]}")
        return codes

def get_lowpoly_streets_by_code(code_str):
    path = os_utils.get_documents_dir()
    
    if ensure_import_openpyxl():
        from openpyxl import load_workbook
        
        file_name = "lp_codes.xlsx"

        logger.add(f"Path to lowpoly codes {os.path.join(path, file_name)}")

        workbook = load_workbook(filename=os.path.join(path, file_name))
        sheet = workbook.active
        streets = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if code_str in str(row[6]):
                streets.append(row[1])
        return streets

def get_requirements_data(highpoly=False):
    path = os_utils.get_documents_dir()
    
    if ensure_import_openpyxl():
        from openpyxl import load_workbook
        
        file_name = r"hp_requirements_V4.xlsx" if highpoly else r"lp_requirements_V4.xlsx"

        logger.add(f"Path to requirements file {os.path.join(path, file_name)}")

        workbook = load_workbook(filename=os.path.join(path, file_name))
        sheet = workbook.active
        data = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i != 0 and str(row[6]) != "1":
                data.append({"req_id": str(row[0]),
                            "req_num": str(row[1]),
                            "category": str(row[2]),
                            "name": str(row[3]),
                            "description": str(row[4]),
                            "auto": str(row[5]),
                            "recomendation": str(row[6]),
                            "help_link": str(row[7])
                            })
        return data

    # file_name = r"hp_requirements_utf.csv" if highpoly else r"lp_requirements_utf.csv"

    # logger.add(f"Path to requirements file {os.path.join(path, file_name)}")

    # with open(os.path.join(path, file_name), encoding='utf-8-sig') as csvfile:
    #     rows = csv.reader(csvfile, delimiter=';')
        
    #     data = []

    #     for i, row in enumerate(rows):
    #         if i != 0 and row[6] != "1":
    #             data.append({"req_id": row[0],
    #                         "req_num": row[1],
    #                         "category": row[2],
    #                         "name": row[3],
    #                         "description": row[4],
    #                         "auto": row[5],
    #                         "recomendation": row[6],
    #                         "help_link": row[7]
    #                         })

    #     return data

def calculate_all_checks(context, bl_operator, by_collections):
    models_path = bpy.path.abspath(context.scene.agr_scene_properties.path)
    lp_checks, hp_checks, project_data = check_highpoly_lowpoly.run(bl_operator, context.scene.agr_scene_properties.Address, models_path, by_collections)
    context.scene.agr_scene_properties.project_data_address = project_data.address
    if lp_checks:
        context.scene.agr_scene_properties.has_lowpoly = True
    if hp_checks:
        context.scene.agr_scene_properties.has_highpoly = True

    data = [(hp_checks, context.scene.agr_scene_properties.checklist_hp_props.categories, True),
            (lp_checks, context.scene.agr_scene_properties.checklist_lp_props.categories, False)]
    
    for props in data:
        if props[0] == None:
            continue
        checks = props[0]
        categories = props[1]
        highpoly = props[2]
        # UpdateRequrements.update_requrements(highpoly)

        # item.auto off if check raised Exception
        for cat in categories:
            for item in cat.collection:
                # item.check_state = utills.CHECK_STATE_ITEMS[0]
                if item.auto and item.req_num not in checks:
                    item.auto = False
                elif item.req_num in checks:
                    item.auto = True

        # check_nums = []
        # for cat in categories:
        #     for item in cat.collection:
        #         # item.check_state = utills.CHECK_STATE_ITEMS[0]
        #         if item.auto:
        #             check_nums.append(item.req_num)
        
        for key in sorted(checks):
            verified_all = all([ch.verified for ch in checks[key]])
            check_ids = checks[key][0].paragraph_ids
            if verified_all:
                # enable checkbox
                true_counts = [0 for _ in range(len(check_ids))]
                for cat in categories:
                    for item in cat.collection:
                        if item.req_num in check_ids:
                            # check_nums.remove(item.req_num)
                            item.errors_count = sum([len(ch.error_list) for ch in checks[key]])
                            item.check = True
                            item.check_state = utills.CHECK_STATE_ITEMS[1]
                            true_counts[check_ids.index(item.req_num)] += 1
                # logger.add(f"{true_counts}  {check_ids}  {key}")
                continue
            else:
                # add red button and add errors text to list in ui (tooltip or expand list or popup)
                for cat in categories:
                    for item in cat.collection:
                        if item.req_num in check_ids:
                            # check_nums.remove(item.req_num)
                            item.errors_count = sum([len(ch.error_list) for ch in checks[key]])
                            item.check = False
                            item.check_state = utills.CHECK_STATE_ITEMS[2]
                            err_lines = ""
                            for check in checks[key]:
                                if not check.verified:
                                    err_lines += f"    {check.directory}\n"
                                    for err in check.error_list:
                                        err_lines += f"        {err}\n"
                            item.errors_text = err_lines
        # if check_nums:
        #     check_nums_string = ", ".join(check_nums)
        #     logger.add(f"highpoly={highpoly}, len={len(check_nums)} not found requirements - {check_nums_string}")
        # msg = ""
        # for key in checks:
        #     msg += key
        #     verified_all = all([ch.verified for ch in checks[key]])
        #     msg += f"   {verified_all}\n\n"
            # for check in checks[key]:
            #     pass
        ui_utills.generate_text_editor(context)

